from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from django.contrib.auth import logout, authenticate, login
from .permissions import IsManager
from django.views.decorators.csrf import ensure_csrf_cookie
from django.http import JsonResponse
from .models import User, Division, Secteur, Manager, Projet, Supervisor, miseAjour, Rapport , Ressource, RessourceUtilisee
from django.shortcuts import get_object_or_404
from django.contrib import admin
import random
import string
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from rest_framework import status
from django.core.exceptions import ObjectDoesNotExist
from .serializers import ProjetSerializer, MiseAjourSerializer, RessourceSerializer, RessourceUtiliseeSerializer, RapportSerializer
from datetime import datetime
from decimal import Decimal
import json
from datetime import datetime, timedelta
from django.http import HttpResponse
from django.core.files.base import ContentFile
import csv
import io
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

class LogoutView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        logout(request)
        return Response({"detail": "Logged out successfully."})

class LoginView(APIView):
    def post(self, request):
        email = request.data.get("email")
        password = request.data.get("password")
        username = get_object_or_404(User, email=email).username
        print(username)
        user = authenticate(request, username=username, password=password)
        print(user)
        if user is not None:

            login(request, user)
            
            user_data = {
                "role": user.role,
                "email": user.email,
                "username": user.username,
                "first_name": user.first_name,
                "last_name": user.last_name,
                "id": user.id,
                "localisation": user.localisation if user.localisation else None,
                "poste": user.poste if user.poste else None,
                "biography": user.biography if user.biography else None,
                "phone": user.phone if user.phone else None,
                "date_joined": user.date_joined,
            }

            if user.role == "supervisor" and hasattr(user, "supervisor_profile"):
                supervisor = user.supervisor_profile
                user_data["division"] = supervisor.division.name if supervisor.division else None

            if user.role == "manager" and hasattr(user, "manager_profile"):
                manager = user.manager_profile
                user_data["division"] = manager.division.name if manager.division else None
                user_data["secteur"] = manager.secteur.name if manager.secteur else None

            return Response(user_data)
        else:
            return Response({"detail": "Invalid credentials."}, status=401)


@ensure_csrf_cookie
def get_csrf_token(request):
    return JsonResponse({'detail': 'CSRF cookie set'})


class DivisionsAndSecteursView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        divisions = Division.objects.all().values('id', 'name') # type: ignore
        secteurs = Secteur.objects.all().values('id', 'name', 'division_id') # type: ignore
        return Response({
            'divisions': list(divisions),
            'secteurs': list(secteurs)
        })


class ManagersListView(APIView):
    def get(self, request):
        managers = Manager.objects.select_related('user').all() # type: ignore
        
        data = [
            {
                "id": manager.id,
                "user_id": manager.user.id,
                "username": manager.user.username,
                "first_name": manager.user.first_name,
                "last_name": manager.user.last_name,
                "email": manager.user.email,
                "division": manager.division.name if manager.division else None,
                "secteur": manager.secteur.name if manager.secteur else None,
                "numProj": Projet.objects.filter(manager=manager.id).count() # type: ignore
            }
            for manager in managers
        ]
        return Response(data)


class UsersListView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        users = User.objects.all()
        data = [
            {
                "id": user.id,
                "full_name": f"{user.first_name} {user.last_name}",
                "role": user.role,
                "email": user.email,
            }
            for user in users
        ]
        return Response(data)

def generatePassword():
    return "".join(random.choices(string.ascii_letters + string.digits, k=10))

def send_welcome_email(to_email, full_name, password, login_url):
    subject = "Bienvenue sur DSI Manager"
    from_email = 'DSI Manager <saad989011@gmail.com>'  
    to = [to_email]

    # Render the HTML template with context
    html_content = render_to_string(
        "welcomeMail.html",
        {"full_name": full_name, "password": password, "login_url": login_url, "email": to_email}
    )
    text_content = f"Bonjour {full_name},\nVotre mot de passe : {password}\nSe connecter : {login_url}"

    msg = EmailMultiAlternatives(subject, text_content, from_email, to)
    msg.attach_alternative(html_content, "text/html")
    msg.send()

class CreateUserView(APIView):
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        # 1. Generate the password first
        raw_password = generatePassword()
        # 2. Create the user with the raw password
        user = User.objects.create_user(
            first_name=request.data.get("full_name").split(" ")[0],
            last_name=request.data.get("full_name").split(" ")[1],
            username=request.data.get("full_name").replace(" ", "").lower(),
            email=request.data.get("email"),
            password=raw_password,
            role=request.data.get("role"),
        )
        full_name = f"{user.first_name} {user.last_name}"
        # 3. Send the email with the raw password
        send_welcome_email(user.email, full_name, raw_password, "http://localhost:5173/login")
        return Response({"detail": "User created successfully."})


class DeleteUserView(APIView):
    permission_classes = [IsAuthenticated]

    def delete(self, request, user_id):
        try:
            user = User.objects.get(id=user_id)
            user.delete()
            return Response({"detail": "User deleted successfully."}, status=status.HTTP_204_NO_CONTENT)
        except ObjectDoesNotExist:
            return Response({"detail": "User not found."}, status=status.HTTP_404_NOT_FOUND)


class UpdateUserView(APIView):
    permission_classes = [IsAuthenticated]

    def put(self, request, user_id):
        try:
            user = User.objects.get(id=user_id)
            # Update fields individually
            full_name = request.data.get("full_name", "")
            if full_name:
                parts = full_name.split(" ", 1)
                user.first_name = parts[0]
                user.last_name = parts[1] if len(parts) > 1 else ""
            user.email = request.data.get("email", user.email)
            user.role = request.data.get("role", user.role)
            user.save()
            return Response({"detail": "User updated successfully."}, status=status.HTTP_200_OK)
        except ObjectDoesNotExist:
            return Response({"detail": "User not found."}, status=status.HTTP_404_NOT_FOUND)

class UpdateUserPasswordView(APIView):
    permission_classes = [IsAuthenticated]

    def put(self, request, user_id):
        try:
            user = User.objects.get(id=user_id)
            current_password = request.data.get("current_password")
            new_password = request.data.get("password")

            if not user.check_password(current_password):
                return Response({"detail": "Current password is incorrect."}, status=status.HTTP_400_BAD_REQUEST)

            user.set_password(new_password)
            user.save()
            return Response({"detail": "Password updated successfully."}, status=status.HTTP_200_OK)
        except ObjectDoesNotExist:
            return Response({"detail": "User not found."}, status=status.HTTP_404_NOT_FOUND)


class ProjectsListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        projets = Projet.objects.all()
        
        for projet in projets:
            # Convert Decimal to float
            budget = float(projet.budget) if projet.budget is not None else 0
            budget_used = float(projet.budget_used) if projet.budget_used is not None else 0

            # Calculate budget progress
            budget_progress = (budget_used / budget) * 100 if budget and budget_used else 0

            # Calculate time progress
            time_progress = 0
            if projet.start_date and projet.end_date:
                try:
                    start = projet.start_date if isinstance(projet.start_date, datetime) else datetime.strptime(str(projet.start_date), "%Y-%m-%d")
                    end = projet.end_date if isinstance(projet.end_date, datetime) else datetime.strptime(str(projet.end_date), "%Y-%m-%d")
                    now = datetime.now()
                    if end > start:
                        time_progress = ((now - start).total_seconds() / (end - start).total_seconds()) * 100
                        time_progress = max(0, min(100, time_progress))
                except Exception:
                    time_progress = 0
            progress = round((budget_progress + time_progress) / 2)
            
            # Update status based on progress ONLY if not manually set to 'termine'
            old_status = projet.status
            old_progress = getattr(projet, 'progress', None)
            
            # Only auto-update status if it's not manually set to 'termine'
            if projet.status != 'termine':
                if progress == 0:
                    projet.status = 'planifie'
                elif 0 < progress < 100:
                    projet.status = 'en_cours'
                elif progress >= 100:
                    projet.status = 'termine'
                    
            # Check for late projects (only if not manually completed)
            if projet.status != 'termine' and projet.end_date:
                end = projet.end_date if isinstance(projet.end_date, datetime) else datetime.strptime(str(projet.end_date), "%Y-%m-%d")
                if datetime.now() > end and progress < 100:
                    projet.status = 'en_retard'
                    
                    
            projet.progress = progress
            # Save only if changed
            if projet.status != old_status or projet.progress != old_progress:
                projet.save()
        serializer = ProjetSerializer(projets, many=True)
        data = serializer.data
        for i, projet in enumerate(projets):
            data[i]['progress'] = projet.progress
        return Response(data)

class UpdateProjectView(APIView):
    permission_classes = [IsAuthenticated]

    def put(self, request, project_id):
        project = get_object_or_404(Projet, id=project_id)
        project.name = request.data.get("name")
        project.category = request.data.get("category")
        division = Division.objects.get(id=request.data.get("division"))
        project.division = division
        secteur = Secteur.objects.get(name=request.data.get("secteur"))
        project.secteur = secteur
        project.description = request.data.get("description")
        project.budget = request.data.get("budget")
        project.budget_used = request.data.get("budget_used")
        project.start_date = request.data.get("date_debut")
        project.end_date = request.data.get("date_fin")
        project.save()
        return Response({"detail": "Project updated successfully."}, status=status.HTTP_200_OK)


def send_project_assignment_email(to_email, full_name, project_name, start_date, end_date, assigned_by , description):

    subject = "Nouveau projet assigné"
    from_email = 'DSI Manager <saad989011@gmail.com>'
    to = [to_email]

    html_content = render_to_string(
        "projectAssignmentMail.html",
        {"user_fullname": full_name, "project_name": project_name, "start_date": start_date, "end_date": end_date, "assigned_by": assigned_by, "description": description}
    )

    msg = EmailMultiAlternatives(subject, "Bonjour " + full_name + ",\nVous avez été assigné au projet suivant : " + project_name + ",\nDate de début : " + start_date + ",\nDate de fin : " + end_date + ",\nAssigné par : " + assigned_by, from_email, to)
    msg.attach_alternative(html_content, "text/html")
    msg.send()


class CreateProjectView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        print(request.data)
        user = get_object_or_404(User, id=request.user.id)
        supervisor = get_object_or_404(Supervisor, user=user)
        manager_user = get_object_or_404(User, id=request.data.get("manager"))
        manager = get_object_or_404(Manager, user=manager_user)
        division = get_object_or_404(Division, id=request.data.get("division"))
        secteur = get_object_or_404(Secteur, name=request.data.get("secteur"))
        supervisor_full_name = f"{supervisor.user.first_name} {supervisor.user.last_name}"
        full_name = f"{manager.user.first_name} {manager.user.last_name}"
        project = Projet.objects.create( # type: ignore
            supervisor=supervisor,
            name=request.data.get("name"),
            category=request.data.get("category"),
            division=division,
            secteur=secteur,
            description=request.data.get("description"),
            manager=manager,
            budget=request.data.get("budget"),
            start_date=request.data.get("date_debut"),
            end_date=request.data.get("date_fin"),
        )
        
        send_project_assignment_email(manager.user.email, full_name, project.name, project.start_date, project.end_date, supervisor_full_name, project.description)
        return Response({"detail": "Project created successfully."}, status=status.HTTP_201_CREATED)


class deleteProjectView(APIView):
    permission_classes = [IsAuthenticated]

    def delete(self, request, project_id):
        try:
            project = Projet.objects.get(id=project_id)
            project.delete()
            return Response({"detail": "Project deleted successfully."}, status=status.HTTP_204_NO_CONTENT)
        except ObjectDoesNotExist:
            return Response({"detail": "Project not found."}, status=status.HTTP_404_NOT_FOUND)

class AddMiseAjourView(APIView):
    def post(self, request):
        print("Request data:", request.data)

        manager_id = request.data.get('manager')
        project_id = request.data.get('project')
        content = request.data.get('content')
        
        # Validate required fields
        if not all([manager_id, project_id, content]):
            return Response({'detail': 'Tous les champs sont requis.'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # Try to get the user first
            try:
                user = User.objects.get(id=manager_id)
                
            except User.DoesNotExist:
                print("Utilisateur non trouvé")
                
            
            # Try to get the manager profile
            try:
                manager_object = Manager.objects.get(user=user)
            except Manager.DoesNotExist:
                print("manager non trouve")
            
            # Try to get the project
            try:
                project_object = Projet.objects.get(id=project_id)
            except Projet.DoesNotExist:
                print("proet non trouve")
            
            # Create the mise à jour
            miseAjour.objects.create(
                manager=manager_object,
                project=project_object,
                content=content
            )
            return Response({'detail': 'Mise à jour créée avec succès.'}, status=status.HTTP_201_CREATED)

        except Exception as e:
            print(f"Error creating mise à jour: {e}")
            return Response(
                {'detail': f'Erreur lors de la création de la mise à jour: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class GetMiseAjourView(APIView):
    def get(self, request):
        try:
            miseAjour_object = miseAjour.objects.all()
            serializer = MiseAjourSerializer(miseAjour_object, many=True)
            return Response(serializer.data)
        except Exception as e:
            print(f"Error fetching mise à jour: {e}")
            return Response({'detail': 'Erreur lors de la récupération des mises à jour.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class markProjetAsCompleted(APIView): 
    def put(self, request, project_id):
        print(f"markProjetAsCompleted called with project_id: {project_id}")
        print(f"Request method: {request.method}")
        print(f"Request user: {request.user}")
        
        project = get_object_or_404(Projet, id=project_id)
        print(f"Found project: {project.name} (ID: {project.id})")
        print(f"Current status: {project.status}")
        
        project.status = "termine"
        
        project.save()
        print(f"Project {project.name} has been marked as completed. New status: {project.status}")
        
        return Response({'detail': 'Projet marqué comme terminé.'}, status=status.HTTP_200_OK)



class RegisterRessourceView(APIView):
    def post(self, request):
        ressource = Ressource.objects.create(
            name=request.data.get("name"),
            type=request.data.get("type"),
            quantity=request.data.get("quantity")
        )
        return Response({'detail': 'Ressource enregistrée avec succès.'}, status=status.HTTP_201_CREATED)

class RessourcesListView(APIView):
    def get(self, request):
        ressources = Ressource.objects.all()
        serializer = RessourceSerializer(ressources, many=True)
        return Response(serializer.data)

class DeleteRessourceView(APIView):
    def delete(self, request, ressource_id):
        ressource = get_object_or_404(Ressource, id=ressource_id)
        ressource.delete()
        return Response({'detail': 'Ressource supprimée avec succès.'}, status=status.HTTP_204_NO_CONTENT)

class getRessourceUtiliseeView(APIView):
    def get(self, request):
        ressource_utilisee = RessourceUtilisee.objects.all()
        serializer = RessourceUtiliseeSerializer(ressource_utilisee, many=True)
        return Response(serializer.data)

class ModifyRessourceView(APIView):
    def put(self, request, ressource_id):
        ressource = get_object_or_404(Ressource, id=ressource_id)
        ressource.name = request.data.get("name")
        ressource.type = request.data.get("type")
        ressource.quantity = request.data.get("quantity")
        ressource.save()
        return Response({'detail': 'Ressource modifiée avec succès.'}, status=status.HTTP_200_OK)


class AddRessourceUtiliseeView(APIView):
    def post(self, request):
        ressource_id = request.data.get("ressource")
        project_id = request.data.get("project")
        quantity = request.data.get("quantity") 
        ressource = get_object_or_404(Ressource, id=ressource_id)
        project = get_object_or_404(Projet, id=project_id)
        ressource_utilisee = RessourceUtilisee.objects.create(
            ressource=ressource,
            project=project,
            quantity=request.data.get("quantity")
        )
        return Response({'detail': 'Ressource utilisée ajoutée avec succès.'}, status=status.HTTP_201_CREATED)

class DeleteRessourceUtiliseeView(APIView):
    def delete(self, request, ressource_utilisee_id):
        ressource_utilisee = get_object_or_404(RessourceUtilisee, id=ressource_utilisee_id)
        ressource_utilisee.delete()
        return Response({'detail': 'Ressource utilisée supprimée avec succès.'}, status=status.HTTP_204_NO_CONTENT)

class ModifyRessourceUtiliseeView(APIView):
    def put(self, request, ressource_utilisee_id):
        ressource_utilisee = get_object_or_404(RessourceUtilisee, id=ressource_utilisee_id)
        ressource_utilisee.quantity = request.data.get("quantity")
        ressource_utilisee.save()
        return Response({'detail': 'Ressource utilisée modifiée avec succès.'}, status=status.HTTP_200_OK)


class GenerateRapportView(APIView):
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        try:
            # Get parameters from request
            periode = request.data.get('periode', '')
            data_types = request.data.get('data_types', [])
            format_type = request.data.get('format', 'xlsx')
            
            # Create a more meaningful filename
            data_types_str = '_'.join(data_types) if data_types else 'all'
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            
            # Generate report title with more meaningful name
            title = f"Rapport_{data_types_str}_{periode}_{timestamp}"
            
            # Create rapport record
            rapport = Rapport.objects.create(
                title=title,
                periode=periode,
                requested_by=request.user
            )
            
            # Generate report content based on data types
            report_data = {}
            
            if 'projects' in data_types:
                projects = Projet.objects.all()
                report_data['projects'] = []
                for project in projects:
                    report_data['projects'].append({
                        'id': project.id,
                        'name': project.name,
                        'category': project.category,
                        'description': project.description,
                        'start_date': project.start_date,
                        'end_date': project.end_date,
                        'status': project.status,
                        'budget': float(project.budget) if project.budget else 0,
                        'budget_used': float(project.budget_used) if project.budget_used else 0,
                        'division': project.division.name if project.division else None,
                        'secteur': project.secteur.name if project.secteur else None,
                        'manager': f"{project.manager.user.first_name} {project.manager.user.last_name}" if project.manager else None,
                        'supervisor': f"{project.supervisor.user.first_name} {project.supervisor.user.last_name}" if project.supervisor else None,
                    })
            
            if 'resources' in data_types:
                ressources = Ressource.objects.all()
                report_data['resources'] = []
                for ressource in ressources:
                    report_data['resources'].append({
                        'id': ressource.id,
                        'name': ressource.name,
                        'type': ressource.type,
                        'quantity': ressource.quantity,
                    })
                
                ressources_utilisees = RessourceUtilisee.objects.all()
                report_data['resources_used'] = []
                for ressource_utilisee in ressources_utilisees:
                    report_data['resources_used'].append({
                        'id': ressource_utilisee.id,
                        'ressource_name': ressource_utilisee.ressource.name if ressource_utilisee.ressource else None,
                        'project_name': ressource_utilisee.project.name if ressource_utilisee.project else None,
                        'quantity': ressource_utilisee.quantity,
                    })
            
            # Generate file based on format
            if format_type == 'xlsx':
                file_content = self._generate_excel_report(report_data, title)
                file_extension = 'xlsx'
            elif format_type == 'csv':
                file_content = self._generate_csv_report(report_data, title)
                file_extension = 'csv'
            elif format_type == 'pdf':
                file_content = self._generate_pdf_report(report_data, title)
                file_extension = 'pdf'
            else:
                return Response({'detail': 'Format non supporté.'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Save file to rapport
            filename = f"{title}.{file_extension}"
            rapport.file.save(filename, ContentFile(file_content), save=True)
            
            return Response({
                'detail': 'Rapport généré avec succès.',
                'rapport_id': rapport.id,
                'filename': filename
            }, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            return Response(
                {'detail': f'Erreur lors de la génération du rapport: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def _generate_excel_report(self, data, title):
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        if 'projects' in data:
            ws_projects = wb.create_sheet("Projets")
            headers = ['ID', 'Nom', 'Catégorie', 'Description', 'Date début', 'Date fin', 'Statut', 'Budget', 'Budget utilisé', 'Division', 'Secteur', 'Manager', 'Supervisor']
            ws_projects.append(headers)
            
            for project in data['projects']:
                ws_projects.append([
                    project['id'],
                    project['name'],
                    project['category'],
                    project['description'],
                    project['start_date'],
                    project['end_date'],
                    project['status'],
                    project['budget'],
                    project['budget_used'],
                    project['division'],
                    project['secteur'],
                    project['manager'],
                    project['supervisor'],
                ])
        
        if 'resources' in data:
            ws_resources = wb.create_sheet("Ressources")
            headers = ['ID', 'Nom', 'Type', 'Quantité']
            ws_resources.append(headers)
            
            for resource in data['resources']:
                ws_resources.append([
                    resource['id'],
                    resource['name'],
                    resource['type'],
                    resource['quantity'],
                ])
        
        if 'resources_used' in data:
            ws_resources_used = wb.create_sheet("Ressources Utilisées")
            headers = ['ID', 'Ressource', 'Projet', 'Quantité']
            ws_resources_used.append(headers)
            
            for resource_used in data['resources_used']:
                ws_resources_used.append([
                    resource_used['id'],
                    resource_used['ressource_name'],
                    resource_used['project_name'],
                    resource_used['quantity'],
                ])
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
    
    def _generate_csv_report(self, data, title):
        output = io.StringIO()
        writer = csv.writer(output)
        
        if 'projects' in data:
            writer.writerow(['=== PROJETS ==='])
            headers = ['ID', 'Nom', 'Catégorie', 'Description', 'Date début', 'Date fin', 'Statut', 'Budget', 'Budget utilisé', 'Division', 'Secteur', 'Manager', 'Supervisor']
            writer.writerow(headers)
            
            for project in data['projects']:
                writer.writerow([
                    project['id'],
                    project['name'],
                    project['category'],
                    project['description'],
                    project['start_date'],
                    project['end_date'],
                    project['status'],
                    project['budget'],
                    project['budget_used'],
                    project['division'],
                    project['secteur'],
                    project['manager'],
                    project['supervisor'],
                ])
            writer.writerow([])
        
        if 'resources' in data:
            writer.writerow(['=== RESSOURCES ==='])
            headers = ['ID', 'Nom', 'Type', 'Quantité']
            writer.writerow(headers)
            
            for resource in data['resources']:
                writer.writerow([
                    resource['id'],
                    resource['name'],
                    resource['type'],
                    resource['quantity'],
                ])
            writer.writerow([])
        
        if 'resources_used' in data:
            writer.writerow(['=== RESSOURCES UTILISÉES ==='])
            headers = ['ID', 'Ressource', 'Projet', 'Quantité']
            writer.writerow(headers)
            
            for resource_used in data['resources_used']:
                writer.writerow([
                    resource_used['id'],
                    resource_used['ressource_name'],
                    resource_used['project_name'],
                    resource_used['quantity'],
                ])
        
        return output.getvalue().encode('utf-8')
    
    def _generate_pdf_report(self, data, title):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        
        # Title
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 20))
        
        if 'projects' in data:
            elements.append(Paragraph("PROJETS", styles['Heading2']))
            elements.append(Spacer(1, 12))
            
            # Projects table
            project_data = [['ID', 'Nom', 'Statut', 'Budget', 'Manager']]
            for project in data['projects']:
                project_data.append([
                    str(project['id']),
                    project['name'],
                    project['status'],
                    str(project['budget']),
                    project['manager'] or 'N/A'
                ])
            
            project_table = Table(project_data)
            project_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(project_table)
            elements.append(Spacer(1, 20))
        
        if 'resources' in data:
            elements.append(Paragraph("RESSOURCES", styles['Heading2']))
            elements.append(Spacer(1, 12))
            
            # Resources table
            resource_data = [['ID', 'Nom', 'Type', 'Quantité']]
            for resource in data['resources']:
                resource_data.append([
                    str(resource['id']),
                    resource['name'],
                    resource['type'],
                    str(resource['quantity'])
                ])
            
            resource_table = Table(resource_data)
            resource_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(resource_table)
        
        doc.build(elements)
        return buffer.getvalue()


class DownloadRapportView(APIView):
    permission_classes = [IsAuthenticated]
    
    def get(self, request, rapport_id):
        try:
            rapport = get_object_or_404(Rapport, id=rapport_id)
            
            if not rapport.file:
                return Response({'detail': 'Fichier non trouvé.'}, status=status.HTTP_404_NOT_FOUND)
            
            # Determine content type based on file extension
            file_extension = rapport.file.name.split('.')[-1].lower()
            content_type_map = {
                'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'csv': 'text/csv',
                'pdf': 'application/pdf'
            }
            content_type = content_type_map.get(file_extension, 'application/octet-stream')
            
            response = HttpResponse(rapport.file.read(), content_type=content_type)
            response['Content-Disposition'] = f'attachment; filename="{rapport.file.name}"'
            return response
            
        except Exception as e:
            return Response(
                {'detail': f'Erreur lors du téléchargement: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class GetRapportsHistoryView(APIView):
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        try:
            rapports = Rapport.objects.all().order_by('-created_at')
            serializer = RapportSerializer(rapports, many=True)
            
            # Transform data to match frontend expectations
            data = []
            for rapport in serializer.data:
                # Determine format from file extension
                format_type = 'Excel'
                if rapport['file']:
                    file_extension = rapport['file'].split('.')[-1].lower()
                    if file_extension == 'pdf':
                        format_type = 'PDF'
                    elif file_extension == 'csv':
                        format_type = 'CSV'
                    elif file_extension == 'xlsx':
                        format_type = 'Excel'
                
                # Create a more readable display name
                display_name = rapport['title']
                if '_' in display_name:
                    # Remove timestamp from display name
                    parts = display_name.split('_')
                    if len(parts) >= 4:
                        # Format: Rapport_data_types_periode_timestamp
                        data_types = parts[1].replace('_', ', ')
                        periode = parts[2]
                        display_name = f"Rapport {data_types} - {periode}"
                
                data.append({
                    'id': rapport['id'],
                    'name': display_name,
                    'format': format_type,
                    'periode': rapport['periode'] or 'N/A',
                    'createdAt': rapport['created_at']
                })
            
            return Response(data)
            
        except Exception as e:
            return Response(
                {'detail': f'Erreur lors de la récupération des rapports: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )








